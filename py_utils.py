import cv2
import torch
from torch.nn import Module
from torch.utils.data import Dataset
import numpy as np
import logging
from openpyxl import load_workbook, Workbook
import datetime
import os


def log_and_print(*args):
    message = ' '.join(map(str, args))
    rank = int(os.environ.get('RANK', 0))
    if rank == 0:
        print(message)
        logging.info(message)


class random_mini_batches_standardtwoModality(Dataset):
    def __init__(self, x1_data, x2_data, y_data):
        self.x1_data = x1_data
        self.x2_data = x2_data
        self.y_data = y_data

    def __len__(self):
        return len(self.y_data)

    def __getitem__(self, index):
        x1 = self.x1_data[index]
        x2 = self.x2_data[index]
        y = self.y_data[index]
        return x1, x2, y


def same_seeds(seed):
    torch.manual_seed(seed)
    if torch.cuda.is_available():
        torch.cuda.manual_seed(seed)
        torch.cuda.manual_seed_all(seed)
    np.random.seed(seed)
    torch.backends.cudnn.benchmark = False
    torch.backends.cudnn.deterministic = True


def convert_to_one_hot(Y, C):
    # 确保 Y 是整数类型
    Y = Y.astype(np.int64)
    # 独热编码
    Y = np.eye(C)[Y.reshape(-1)].T
    return Y


def generate_batch(idx, X_PCAMirrow, Y, batch_size, patch_size, row, col, num_class, shuffle=False, only_valid = False):
    num = len(idx)
    if shuffle:
        np.random.shuffle(idx)

    for i in range(0, num, batch_size):
        bi = np.array(idx)[np.arange(i, min(num, i + batch_size))]
        index_row = np.ceil((bi + 1) * 1.0 / col).astype(np.int32)
        index_col = (bi + 1) - (index_row - 1) * col

        if only_valid:
            valid_mask = (Y[bi] != 0) & (Y[bi] != -1)
            bi = bi[valid_mask]
            index_row = index_row[valid_mask]
            index_col = index_col[valid_mask]
            if bi.size == 0:
                continue

        # 根据 X_PCAMirrow 的维度来处理
        if X_PCAMirrow.ndim == 2:
            # 2维数据 (如 LiDAR)
            patches = np.zeros([bi.size, patch_size * patch_size])
            for j in range(bi.size):
                a = index_row[j] - 1
                b = index_col[j] - 1
                patch = X_PCAMirrow[a:a + patch_size, b:b + patch_size]
                patches[j, :] = patch.reshape(patch_size * patch_size)
        else:
            # 3维数据 (如 HSI)
            patches = np.zeros([bi.size, patch_size * patch_size * X_PCAMirrow.shape[-1]])
            for j in range(bi.size):
                a = index_row[j] - 1
                b = index_col[j] - 1
                patch = X_PCAMirrow[a:a + patch_size, b:b + patch_size, :]
                patches[j, :] = patch.reshape(patch_size * patch_size * X_PCAMirrow.shape[-1])

        yield patches, bi


def sampling(Y_train, Y_test):
    n_class = int(Y_test.max())  # 转换为整数
    train_idx = list()
    test_idx = list()

    for i in range(1, n_class + 1):
        train_i = np.where(Y_train == i)[0]
        test_i = np.where(Y_test == i)[0]

        train_idx.extend(train_i)
        test_idx.extend(test_i)

    train_idx = np.array(train_idx)
    test_idx = np.array(test_idx)

    return train_idx, test_idx


def generate_cube(idx, X, Y, patch_size, row, col, num_class, shuffle=False, augment=True):
    num_class = int(num_class)  # 确保是整数
    if shuffle:
        np.random.shuffle(idx)
    bi = np.array(idx)
    index_row = np.ceil((bi + 1) / col).astype(np.int32)
    index_col = (bi + 1) - (index_row - 1) * col

    # 检查 X 的维度
    if X.ndim == 2:
        # 2维数据 (如 LiDAR 标准化后)
        patches = np.zeros([bi.size, patch_size * patch_size])
        for j in range(bi.size):
            a = index_row[j] - 1
            b = index_col[j] - 1
            patch = X[a:a + patch_size, b:b + patch_size]
            patches[j, :] = patch.reshape(patch_size * patch_size)
    else:
        # 3维数据 (如 HSI)
        patches = np.zeros([bi.size, patch_size * patch_size * X.shape[-1]])
        for j in range(bi.size):
            a = index_row[j] - 1
            b = index_col[j] - 1
            patch = X[a:a + patch_size, b:b + patch_size, :]
            patches[j, :] = patch.reshape(patch_size * patch_size * X.shape[-1])

    labels = Y[bi] - 1
    labels = convert_to_one_hot(labels, num_class)
    labels = labels.T

    # 在返回前添加数据增强
    if augment and np.random.random() > 0.5:
        # 随机翻转
        patches = patches.reshape(-1, patch_size, patch_size, X.shape[-1])
        if np.random.random() > 0.5:
            patches = np.flip(patches, axis=1)  # 上下翻转
        if np.random.random() > 0.5:
            patches = np.flip(patches, axis=2)  # 左右翻转
        patches = patches.reshape(-1, patch_size * patch_size * X.shape[-1])

    return patches, labels


def create_excel(excel_path, mode, labels):
    try:
        workbook = load_workbook(excel_path)
    except FileNotFoundError:
        workbook = Workbook()

    if mode not in workbook.sheetnames:
        worksheet = workbook.create_sheet(mode)
    else:
        worksheet = workbook[mode]

    for col, header in enumerate(['Time', 'epoch', 'svd'] + labels + ['OA', 'AA', 'K'], start=1):
        worksheet.cell(row=1, column=col).value = header

    return workbook, worksheet


def mapset_normalization(mapset, type, concate_pixel, n_feature=0):
    print(type, '.shape = ', mapset.shape)
    expand = (mapset.ndim == 2) and (n_feature != 0)
    if expand == True:
        mapset = mapset.reshape(mapset.shape[0], mapset.shape[1], 1)
    [row, col, feature] = mapset.shape
    mapset = mapset.reshape(row * col, feature)
    mapset = np.asarray(mapset, dtype=np.float32)
    mapset = (mapset - np.min(mapset)) / (np.max(mapset) - np.min(mapset))
    mapset = mapset.reshape(row, col, feature)
    mapset = cv2.copyMakeBorder(mapset, top=concate_pixel[0], bottom=concate_pixel[1], left=concate_pixel[2],
                                right=concate_pixel[3], borderType=cv2.BORDER_REFLECT)
    if expand == True:
        mapset = mapset.reshape(row + concate_pixel[2] + concate_pixel[3], col + concate_pixel[0] + concate_pixel[1], 1)
        mapset = np.repeat(mapset, n_feature, axis=-1)
    print(type, '.shape(after normalization) = ', mapset.shape)
    return mapset, row, col, n_feature


def result_cal(c):
    overall_acc = torch.sum(torch.diag(c)) / torch.sum(c)
    per_class_acc = torch.diag(c) / torch.sum(c, axis=1)
    true_total = torch.sum(c, axis=1)
    pre_toral = torch.sum(c, axis=0)
    sample_total = torch.sum(c)
    change_agreement = torch.sum(true_total * pre_toral) / (sample_total * sample_total)
    k = (overall_acc - change_agreement) / (1 - change_agreement)
    aa = sum(per_class_acc) / len(per_class_acc)
    return overall_acc, per_class_acc, aa, k


def save_result_excel(workbook, worksheet, excel_path, epoch, svd_k, overall_acc, per_class_acc, aa, k):
    '''
    save result to excel
    '''
    data = []
    for count, number in enumerate(per_class_acc, start=1):
        log_and_print("This is round {0}, the accuracy of class {1}: {2:.4f}".format(epoch, count, number))
        data.append(number * 100)
    data.append(overall_acc * 100)
    data.append(aa * 100)
    data.append(k)

    new_row = worksheet.max_row + 1
    worksheet.cell(row=new_row, column=1).value = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    worksheet.cell(row=new_row, column=2).value = "{:d}".format(epoch)
    worksheet.cell(row=new_row, column=3).value = "{:d}".format(svd_k)
    for i in range(0, len(data)):
        worksheet.cell(row=new_row, column=3 + (i + 1)).value = "{:.2f}".format(data[i])
    workbook.save(excel_path)